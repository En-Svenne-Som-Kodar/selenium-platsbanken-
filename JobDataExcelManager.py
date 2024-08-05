import openpyxl
from openpyxl import load_workbook
from datetime import datetime


class JobDataExcelManager:
    def __init__(self, file_name):
        self.file_name = file_name
        self.workbook, self.sheet = self.initialize_workbook()

    def initialize_workbook(self):
        try:
            workbook = load_workbook(self.file_name)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["URL", "Title", "Företag", "Arbetsroll", "Kommun", "Annons ID", "Datum"])

        return workbook, sheet

    def add_data(self, data):
        existing_ad_ids = [row[5] for row in self.sheet.iter_rows(min_row=2, max_col=6,
                                                                  values_only=True)]  # Annons ID är på den sjätte kolumnen (index 5)

        if data[5] not in existing_ad_ids:
            self.sheet.append(data)
        else:
            print(f"Annons ID {data[5]} finns redan i filen.")

    def save_workbook(self):
        self.workbook.save(self.file_name)


